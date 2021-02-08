#############################START LICENSE##########################################
# Copyright (C) 2019 Pedro Martinez
#
# # This program is free software: you can redistribute it and/or modify
# # it under the terms of the GNU Affero General Public License as published
# # by the Free Software Foundation, either version 3 of the License, or
# # (at your option) any later version (the "AGPL-3.0+").
#
# # This program is distributed in the hope that it will be useful,
# # but WITHOUT ANY WARRANTY; without even the implied warranty of
# # MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# # GNU Affero General Public License and the additional terms for more
# # details.
#
# # You should have received a copy of the GNU Affero General Public License
# # along with this program. If not, see <http://www.gnu.org/licenses/>.
#
# # ADDITIONAL TERMS are also included as allowed by Section 7 of the GNU
# # Affero General Public License. These additional terms are Sections 1, 5,
# # 6, 7, 8, and 9 from the Apache License, Version 2.0 (the "Apache-2.0")
# # where all references to the definition "License" are instead defined to
# # mean the AGPL-3.0+.
#
# # You should have received a copy of the Apache-2.0 along with this
# # program. If not, see <http://www.apache.org/licenses/LICENSE-2.0>.
#############################END LICENSE##########################################


###########################################################################################
#
#   Script name: qc-ICProfile
#
#   Description: Tool for batch processing and report generation of ICProfile files
#
#   Example usage: python qc-ICProfile "/folder/"
#
#   Author: Pedro Martinez
#   pedro.enrique.83@gmail.com
#   5877000722
#   Date:2019-04-09
#
###########################################################################################



import os
import sys
import pydicom
import re
import argparse
import linecache
import tokenize
from PIL import *
import subprocess
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.backends.backend_pdf import PdfPages
from tqdm import tqdm
import numpy as np
import pandas as pd
from openpyxl import Workbook, cell, load_workbook
from math import *
import scipy.integrate as integrate
from int_detc_indx import int_detc_indx




def area_calc(profile,coord):
    # print(profile,coord)
    area = np.trapz(profile,coord)
    return area






def read_icp(dirname):
# this section reads the header and detects to what cells to write
    
#we need to read the calibration file that corresponds with     
    with os.scandir(dirname) as entries:
        Data_dict = {}
        Data_dict2 = {}
        key_list=[]
        key_list2=[]
        key_list.append('X')
        key_list.append('Y')  
        key_list2.append('X')
        key_list2.append('Y')

        # These vectors will have the location of the sensors in the x, y and diagonal directions
        Y = (np.linspace(1,65,65)-33)/2
        X= np.delete(np.delete(Y,31),32)
        PD = np.delete(np.delete((np.linspace(1,65,65)-33)/2,31),32)
        ND=PD
        PDX = PD/ np.cos(pi / 4)
        PDY = PD/ np.sin(pi / 4)
        NDX = ND/ np.cos(pi / 4 - pi / 2)
        NDY = ND/ np.sin(pi / 4 - pi / 2)

        Data_dict['X']=np.pad(X,(0,2),'constant',constant_values=0)
        Data_dict['Y']=Y
        Data_dict2['X']=np.pad(X[:len(X)//2],(0,1),'constant',constant_values=0)
        Data_dict2['Y']=Y[:len(Y)//2]
         #PD and ND also use the same number of detectors
        for filename in entries:
            if filename.is_file():
                if os.path.splitext(filename.name)[1]=='.prs':
                    print('Start header processing')
                    file = open(filename, mode='r', encoding='ISO-8859-1') #,encoding='utf-8-sig')
                    lines = file.readlines()
                    file.close()
                    calname = 'N:'+lines[5].rstrip().split('N:')[1]
                    # gain = int(re.findall(r'\d+',lines[20])[0])  # using regex
                    gain = int(lines[20].rstrip().split('\t')[1])
                    print(lines[29].split('\t'))
                    mode = lines[29].rstrip().split('\t')[1]
                    energy = lines[29].rstrip().split('\t')[3]
                    print('Calibration file name = ',calname)
                    print('Gain = ',gain)
                    print('Mode = ',mode)
                    print('Energy = ',energy)
                    my_dict = {}
                    my_list = []
                    if mode=='X-Ray FFF' and gain!=2:
                        print('Error, gain was set incorrectly')
                        exit(0)
                    print('Start data processing')
                    # reading measurement file
                    df = pd.read_csv(filename,skiprows=106,delimiter='\t')
                    tblen = df.shape[0]  #length of the table
                    #These vectors will hold the inline and crossline data
                    RawCountXvect = []
                    CorrCountXvect=[] # correcting for leakage using the expression # for Detector(n) = {RawCount(n) - TimeTic * LeakRate(n)} * cf(n)
                    RawCountYvect = []
                    CorrCountYvect=[]
                    RawCountPDvect = [] # positive diagonal
                    CorrCountPDvect=[]
                    RawCountNDvect = [] # negative diagonal
                    CorrCountNDvect=[]
                    BiasX=[]
                    CalibX=[]
                    BiasY=[]
                    CalibY=[]
                    BiasPD=[]
                    CalibPD=[]
                    BiasND=[]
                    CalibND=[]
                    Timetic=df['TIMETIC'][3]

                    QuadWedgeCal=[0.5096,0,0,0,0,0,0,0,0] # 6xqw,15xqw,6fffqw,10fffqw,6eqw,9eqw,12eqw,16eqw,20eqw
                    # figs = [] #in this list we will hold all the figures
                    print('Timetic=',Timetic*1e-6,df['TIMETIC']) # duration of the measurement
                    # print('Backrate',df)
                    for column in df.columns[5:68]: #this section records the X axis (-)
                        CorrCountXvect.append((df[column][3] - Timetic*df[column][0])*df[column][1]/gain)#the corrected data for leakage = Timetic*Bias*Calibration/Gain
                        BiasX.append(df[column][0]) # already used in the formula above but saving them just in case
                        CalibX.append(df[column][1])
                        RawCountXvect.append(df[column][3])
                    for column in df.columns[68:133]: #this section records the Y axis (|)
                        CorrCountYvect.append((df[column][3] - Timetic*df[column][0])*df[column][1]/gain)
                        BiasY.append(df[column][0]) # already used in the formula above but saving them just in case
                        CalibY.append(df[column][1])
                        RawCountYvect.append(df[column][3])
                    for column in df.columns[133:196]: #this section records the D1 axis  (/)
                        CorrCountPDvect.append((df[column][3] - Timetic*df[column][0])*df[column][1]/gain)
                        BiasPD.append(df[column][0]) # already used in the formula above but saving them just in case
                        CalibPD.append(df[column][1])
                        RawCountPDvect.append(df[column][3])
                    for column in df.columns[196:259]: #this section records the D2 axis  (\)
                        CorrCountNDvect.append((df[column][3] - Timetic*df[column][0])*df[column][1]/gain)
                        BiasND.append(df[column][0]) # already used in the formula above but saving them just in case
                        CalibND.append(df[column][1])
                        RawCountNDvect.append(df[column][3])


                    FRGN=80
                    xli,xri, xlFRGN, xrFRGN,CMX = int_detc_indx(CorrCountXvect,FRGN)
                    yli,yri,ylFRGN, yrFRGN,CMY = int_detc_indx(CorrCountYvect,FRGN)
                    pdli,pdri,pdlFRGN, pdrFRGN,CMPD = int_detc_indx(CorrCountPDvect,FRGN)
                    ndli,ndri,ndlFRGN, ndrFRGN,CMND = int_detc_indx(CorrCountNDvect,FRGN)


                    # wb = Workbook()
                    # ws = wb.active
                    # ws.append()
                    # wb.save(dirname+"output.xlsx")
                    print(os.path.splitext(filename.name)[0])
                    key_list.append(os.path.splitext(filename.name)[0]+'_X')
                    key_list.append(os.path.splitext(filename.name)[0]+'_Y')
                    key_list.append(os.path.splitext(filename.name)[0]+'_PD')
                    key_list.append(os.path.splitext(filename.name)[0]+'_ND')

                    key_list2.append(os.path.splitext(filename.name)[0]+'_X')
                    key_list2.append(os.path.splitext(filename.name)[0]+'_Y')
                    key_list2.append(os.path.splitext(filename.name)[0]+'_PD')
                    key_list2.append(os.path.splitext(filename.name)[0]+'_ND')

                    # print(CorrCountXvect)
                    symmetryXVect = (np.flip(CorrCountXvect) - CorrCountXvect)/CorrCountXvect[len(CorrCountXvect)//2]*100
                    # symmetry_X=max(symmetryXVect[xli:len(symmetryXVect)//2],key=abs)
                    # print(xli,symmetry_X)
                    # for i in range(0,len(CorrCountXvect)//2):
                    # #     symDiffXvect.append(CorrCountXvect[i]-CorrCountXvect[-i-1])
                    #     print(i,X[i],CorrCountXvect[i],np.flip(CorrCountXvect)[i],symmetryXVect[i])
                    # #     symDiffPDvect.append(CorrCountPDvect[i]-CorrCountPDvect[-i-1])
                    # #     symDiffNDvect.append(CorrCountNDvect[i]-CorrCountNDvect[-i-1])
                    
                    print(CorrCountYvect)
                    symmetryYVect = (np.flip(CorrCountYvect) - CorrCountYvect)/CorrCountYvect[len(CorrCountYvect)//2]*100
                    
                    symmetry_Y=max(symmetryYVect[yli:len(symmetryYVect)//2],key=abs)
                    for i in range(0,len(CorrCountYvect)//2):
                    #     symDiffXvect.append(CorrCountXvect[i]-CorrCountXvect[-i-1])
                        print(i,Y[i],CorrCountYvect[i],np.flip(CorrCountYvect)[i],symmetryYVect[i])
                    #     symDiffPDvect.append(CorrCountPDvect[i]-CorrCountPDvect[-i-1])
                    #     symDiffNDvect.append(CorrCountNDvect[i]-CorrCountNDvect[-i-1])

                    symmetryPDVect = (np.flip(CorrCountPDvect) - CorrCountPDvect)/CorrCountYvect[len(CorrCountPDvect)//2]*100
                    symmetryNDVect = (np.flip(CorrCountNDvect) - CorrCountNDvect)/CorrCountYvect[len(CorrCountNDvect)//2]*100

                    symmetry_X=max(symmetryXVect[xli:len(symmetryXVect)//2],key=abs)
                    symmetry_Y=max(symmetryYVect[yli:len(symmetryYVect)//2],key=abs)
                    symmetry_PD=max(symmetryPDVect[pdli:len(symmetryPDVect)//2],key=abs)
                    symmetry_ND=max(symmetryNDVect[ndli:len(symmetryNDVect)//2],key=abs)
                    index_sym_X = np.argmax(np.abs(symmetryXVect[xli:len(symmetryXVect)//2]))
                    index_sym_Y = np.argmax(np.abs(symmetryYVect[yli:len(symmetryYVect)//2]))
                    index_sym_PD = np.argmax(np.abs(symmetryPDVect[pdli:len(symmetryPDVect)//2]))
                    index_sym_ND = np.argmax(np.abs(symmetryNDVect[ndli:len(symmetryNDVect)//2]))
                    print(xli,X[xli],'amax(symmXVect)',symmetry_X,index_sym_X,X[xli+index_sym_X])
                    print(yli,Y[yli],'amax(symmYVect)',symmetry_Y,index_sym_Y,Y[yli+index_sym_Y])
                    print(pdli,PD[pdli],'amax(symmPDVect)',symmetry_PD,index_sym_PD,PD[pdli+index_sym_PD])
                    print(ndli,ND[ndli],'amax(symmNDVect)',symmetry_ND,index_sym_ND,ND[ndli+index_sym_ND])

                    
                    #we need to pad these arrays
                    CorrCountXvect.extend([0,0])
                    CorrCountPDvect.extend([0,0])
                    CorrCountNDvect.extend([0,0])


                    symmetryXVect = symmetryXVect.tolist()
                    symmetryPDVect = symmetryPDVect.tolist()
                    symmetryNDVect = symmetryNDVect.tolist()

                    symmetryXVect.extend([0,0])
                    symmetryPDVect.extend([0,0])
                    symmetryNDVect.extend([0,0])


                    print(len(symmetryXVect),len(symmetryYVect),len(symmetryPDVect),len(symmetryNDVect))



                    Data_dict[os.path.splitext(filename.name)[0]+'_X']=CorrCountXvect
                    Data_dict[os.path.splitext(filename.name)[0]+'_Y']=CorrCountYvect
                    Data_dict[os.path.splitext(filename.name)[0]+'_PD']=CorrCountPDvect
                    Data_dict[os.path.splitext(filename.name)[0]+'_ND']=CorrCountNDvect

                    Data_dict2[os.path.splitext(filename.name)[0]+'_X']=symmetryXVect[:len(symmetryXVect)//2]
                    Data_dict2[os.path.splitext(filename.name)[0]+'_Y']=symmetryYVect[:len(symmetryYVect)//2]
                    Data_dict2[os.path.splitext(filename.name)[0]+'_PD']=symmetryPDVect[:len(symmetryPDVect)//2]
                    Data_dict2[os.path.splitext(filename.name)[0]+'_ND']=symmetryNDVect[:len(symmetryNDVect)//2]
                    

        df = pd.DataFrame(Data_dict,columns=key_list)
        df2 = pd.DataFrame(Data_dict2,columns=key_list2)

        
        book = load_workbook(dirname+'curves.xlsx')
        writer = pd.ExcelWriter(dirname+'curves.xlsx',engine='openpyxl')
        writer.book = book

        # print(df.describe,df2.describe)
        df.to_excel(writer,sheet_name='Sheet1')
        df2.to_excel(writer,sheet_name='Sheet2')
        writer.save()
        writer.close()
        exit(0)







if __name__ == "__main__":
    parser = argparse.ArgumentParser()  # pylint: disable = invalid-name
    parser.add_argument( "folder", help="path to folder")
    args = parser.parse_args()  # pylint: disable = invalid-name

    if args.folder:
        dirname = args.folder  # pylint: disable = invalid-name
        read_icp(dirname)












